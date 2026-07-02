"""Strategy pattern for data export.

Provides a common interface for exporting data in different formats,
making it easy to add new formats without modifying existing code.
"""

from __future__ import annotations

import csv
import json
from abc import ABC, abstractmethod
from pathlib import Path

from utils.logger import setup_logger

logger = setup_logger(__name__)


class ExportStrategy(ABC):
    """Abstract base class for export strategies."""

    @abstractmethod
    def export(self, data: list[dict], path: Path) -> Path:
        """Export data to the given path. Returns the actual file path."""

    @abstractmethod
    def get_extension(self) -> str:
        """Return the file extension (e.g., '.csv', '.json')."""


class CSVExport(ExportStrategy):
    """Export data to CSV format."""

    def export(self, data: list[dict], path: Path) -> Path:
        if not data:
            raise ValueError("No data to export")

        filepath = path.with_suffix(self.get_extension())
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

        logger.info("CSV exported: %s", filepath)
        return filepath

    def get_extension(self) -> str:
        return ".csv"


class JSONExport(ExportStrategy):
    """Export data to JSON format."""

    def export(self, data: list[dict], path: Path) -> Path:
        filepath = path.with_suffix(self.get_extension())
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info("JSON exported: %s", filepath)
        return filepath

    def get_extension(self) -> str:
        return ".json"


class PDFExport(ExportStrategy):
    """Export data to PDF format using reportlab."""

    def export(self, data: list[dict], path: Path) -> Path:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except ImportError:
            raise ImportError("reportlab is required for PDF export")

        if not data:
            raise ValueError("No data to export")

        filepath = path.with_suffix(self.get_extension())
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )

        elements = []

        headers = list(data[0].keys())
        table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in data]
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F5F5F5")],
                    ),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)

        logger.info("PDF exported: %s", filepath)
        return filepath

    def get_extension(self) -> str:
        return ".pdf"


class XLSXExport(ExportStrategy):
    """Export data to Excel XLSX format using openpyxl."""

    def export(self, data: list[dict], path: Path) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export")

        if not data:
            raise ValueError("No data to export")

        filepath = path.with_suffix(self.get_extension())
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, row in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=row.get(header, ""))

        wb.save(str(filepath))
        logger.info("XLSX exported: %s", filepath)
        return filepath

    def get_extension(self) -> str:
        return ".xlsx"


# Factory registry
EXPORT_STRATEGIES: dict[str, type[ExportStrategy]] = {
    "csv": CSVExport,
    "json": JSONExport,
    "pdf": PDFExport,
    "xlsx": XLSXExport,
}


def get_export_strategy(fmt: str) -> ExportStrategy:
    """Get an export strategy instance by format name.

    Args:
        fmt: Format name ('csv', 'json', 'pdf', 'xlsx').

    Returns:
        An instance of the corresponding ExportStrategy.

    Raises:
        ValueError: If the format is not supported.
    """
    strategy_cls = EXPORT_STRATEGIES.get(fmt.lower())
    if not strategy_cls:
        raise ValueError(f"Unsupported export format: {fmt}. Use one of: {list(EXPORT_STRATEGIES)}")
    return strategy_cls()
