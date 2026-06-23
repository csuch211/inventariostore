"""
Export service for generating reports in multiple formats
"""

import csv
from collections.abc import Callable
from datetime import datetime
from pathlib import Path

from utils.exceptions import InventarioException
from utils.logger import setup_logger

logger = setup_logger(__name__)


class ExportService:
    """Handle data export to various formats"""

    def __init__(self, export_dir: Path | None = None):
        self.export_dir = export_dir or Path("./exports")
        self.export_dir.mkdir(exist_ok=True)

    def export_to_csv(self, productos: list[dict], filename: str | None = None) -> Path:
        """Export products to CSV"""
        try:
            if not filename:
                filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            filepath = self.export_dir / filename

            if not productos:
                logger.warning("No data to export")
                raise InventarioException("No products to export")

            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=productos[0].keys())
                writer.writeheader()
                writer.writerows(productos)

            logger.info(f"CSV exported: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            raise InventarioException(f"Export failed: {e}")

    def export_to_json(self, productos: list[dict], filename: str | None = None) -> Path:
        """Export products to JSON"""
        try:
            import json

            if not filename:
                filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

            filepath = self.export_dir / filename

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(productos, f, indent=2, ensure_ascii=False)

            logger.info(f"JSON exported: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting JSON: {e}")
            raise InventarioException(f"Export failed: {e}")

    def export_summary_report(self, stats: dict, filename: str | None = None) -> Path:
        """Export summary report"""
        try:
            if not filename:
                filename = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

            filepath = self.export_dir / filename

            content = f"""
╔══════════════════════════════════════╗
║  REPORTE DE INVENTARIO EMPRESARIAL   ║
╚══════════════════════════════════════╝

Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

ESTADÍSTICAS:
─────────────────────────────────────
Total de productos: {stats.get("total_productos", 0)}
Cantidad total: {stats.get("cantidad_total", 0)} unidades
Valor total: ${stats.get("valor_total", 0):,.2f}

──────────────────────────────────────
Generado por: Sistema de Inventario
"""

            with open(filepath, "w", encoding="utf-8") as f:
                f.write(content)

            logger.info(f"Report exported: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting report: {e}")
            raise InventarioException(f"Export failed: {e}")

    def export_to_pdf(
        self,
        productos: list[dict],
        filename: str | None = None,
        title: str = "Reporte de Inventario",
        columns: list[str] | None = None,
    ) -> Path:
        """Export products to PDF using reportlab.

        Args:
            productos: list of dicts (one row per product).
            filename: optional filename; defaults to timestamped.
            title: document title shown at the top.
            columns: optional list of column tuples (header_key, header_label)
                controlling both the rendered headers and the per-row keys.
                When None, the legacy default schema is used.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )

            if not filename:
                filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            filepath = self.export_dir / filename

            if not productos:
                logger.warning("No data to export")
                raise InventarioException("No products to export")

            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=landscape(letter),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30,
            )

            styles = getSampleStyleSheet()
            elements = []

            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Title"],
                fontSize=18,
                spaceAfter=6,
                textColor=colors.HexColor("#1976D2"),
            )
            elements.append(Paragraph(title, title_style))
            elements.append(
                Paragraph(
                    f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 0.25 * inch))

            total_qty = sum(p.get("cantidad", 0) for p in productos)
            total_val = sum(p.get("precio", 0) * p.get("cantidad", 0) for p in productos)
            elements.append(
                Paragraph(
                    f"Productos: {len(productos)} | Unidades totales: {total_qty} | Valor total: ${total_val:,.2f}",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 0.25 * inch))

            if columns is None:
                columns = [
                    ("codigo", "Código"),
                    ("nombre", "Nombre"),
                    ("cantidad", "Cantidad"),
                    ("precio", "Precio"),
                    ("categoria", "Categoría"),
                ]

            def _cell_for(p: dict, key: str) -> str:
                if key == "precio":
                    return f"${p.get('precio', 0):.2f}"
                if key == "stock_min":
                    return str(p.get("stock_min", 0))
                if key == "alert_level":
                    level = p.get("alert_level", "")
                    return (
                        "Sin stock"
                        if level == "critical"
                        else ("Bajo" if level == "low" else level)
                    )
                return str(p.get(key, ""))

            headers = [label for _, label in columns]
            data = [headers]
            for p in productos:
                data.append([_cell_for(p, key) for key, _ in columns])

            # Heuristic column widths based on key
            width_map = {
                "codigo": 1.2 * inch,
                "nombre": 3.5 * inch,
                "cantidad": 1.0 * inch,
                "precio": 1.0 * inch,
                "categoria": 1.5 * inch,
                "stock_min": 1.0 * inch,
                "alert_level": 1.3 * inch,
            }
            col_widths = [width_map.get(key, 1.2 * inch) for key, _ in columns]
            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("ALIGN", (2, 0), (3, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F5F5F5")],
                        ),
                    ]
                )
            )
            elements.append(table)

            doc.build(elements)
            logger.info(f"PDF exported: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}")
            raise InventarioException(f"Export failed: {e}")

    def export_to_xlsx(
        self,
        productos: list[dict],
        filename: str | None = None,
        title: str = "Reporte de Inventario",
        columns: list[tuple] | None = None,
    ) -> Path:
        """Export products to Excel XLSX using openpyxl.

        Args:
            productos: list of dicts.
            filename: optional filename.
            title: title row on the Resumen sheet.
            columns: optional list of (key, label) tuples to drive the main
                sheet. When None, the legacy default schema is used.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            if not filename:
                filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            filepath = self.export_dir / filename

            if not productos:
                logger.warning("No data to export")
                raise InventarioException("No products to export")

            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"

            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            if columns is None:
                columns = [
                    ("codigo", "Código"),
                    ("nombre", "Nombre"),
                    ("cantidad", "Cantidad"),
                    ("precio", "Precio"),
                    ("categoria", "Categoría"),
                    ("descripcion", "Descripción"),
                ]

            for col, (_, header) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            def _xlsx_cell(p: dict, key: str):
                if key == "alert_level":
                    level = p.get("alert_level", "")
                    return (
                        "Sin stock"
                        if level == "critical"
                        else ("Bajo" if level == "low" else level)
                    )
                v = p.get(key, "")
                return v if v is not None else ""

            for row_idx, p in enumerate(productos, 2):
                for col, (key, _) in enumerate(columns, 1):
                    cell = ws.cell(row=row_idx, column=col, value=_xlsx_cell(p, key))
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            # Heuristic column widths
            width_map = {
                "codigo": 14,
                "nombre": 35,
                "cantidad": 10,
                "precio": 12,
                "categoria": 16,
                "descripcion": 30,
                "stock_min": 12,
                "alert_level": 14,
            }
            for col_idx, (key, _) in enumerate(columns, 1):
                letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[letter].width = width_map.get(key, 14)

            summary_ws = wb.create_sheet("Resumen")
            total_qty = sum(p.get("cantidad", 0) for p in productos)
            total_val = sum(p.get("precio", 0) * p.get("cantidad", 0) for p in productos)

            summary_data = [
                (title, ""),
                ("", ""),
                ("Total de productos", len(productos)),
                ("Unidades totales", total_qty),
                ("Valor total", f"${total_val:,.2f}"),
                ("", ""),
                ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]

            for row_idx, (label, value) in enumerate(summary_data, 1):
                cell_lbl = summary_ws.cell(row=row_idx, column=1, value=label)
                summary_ws.cell(row=row_idx, column=2, value=value)
                if row_idx == 1:
                    cell_lbl.font = Font(bold=True, size=14, color="1976D2")
                else:
                    cell_lbl.font = Font(bold=True)

            summary_ws.column_dimensions["A"].width = 22
            summary_ws.column_dimensions["B"].width = 20

            wb.save(str(filepath))
            logger.info(f"XLSX exported: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting XLSX: {e}")
            raise InventarioException(f"Export failed: {e}")

    # ============ CSV Import ============

    @staticmethod
    def import_from_csv(
        filepath: str,
        crear_producto_fn: Callable[..., dict],
    ) -> tuple[int, list[str]]:
        """Import products from a CSV file.

        Expected header (case-insensitive): codigo, nombre, cantidad, precio,
        categoria (optional), descripcion (optional), stock_min (optional).

        Args:
            filepath: Path to the CSV file.
            crear_producto_fn: Callable that receives kwargs (codigo, nombre,
                cantidad, precio, categoria, descripcion, stock_min) and
                returns the created product dict. Wraps DB.crear_producto so
                errors propagate per row.

        Returns:
            Tuple of (success_count, list_of_error_messages_per_row).
        """
        path = Path(filepath)
        if not path.exists():
            raise InventarioException(f"CSV file not found: {filepath}")

        success = 0
        errors: list[str] = []
        try:
            with open(path, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    raise InventarioException("CSV appears to be empty")

                # Normalize field names to lowercase for tolerant matching
                normalized_fields = {f.lower().strip(): f for f in reader.fieldnames}
                required = ("codigo", "nombre")
                for r in required:
                    if r not in normalized_fields:
                        raise InventarioException(f"Missing required column: {r}")

                for row_num, raw_row in enumerate(reader, start=2):  # header is row 1
                    row = {k.lower().strip(): (v or "").strip() for k, v in raw_row.items() if k}
                    try:
                        codigo = row.get("codigo", "")
                        nombre = row.get("nombre", "")
                        if not codigo or not nombre:
                            errors.append(f"Row {row_num}: codigo and nombre are required")
                            continue
                        try:
                            cantidad = int(row.get("cantidad", "0") or "0")
                            precio = float(row.get("precio", "0") or "0")
                            stock_min = int(row.get("stock_min", "0") or "0")
                        except ValueError:
                            errors.append(
                                f"Row {row_num}: cantidad/precio/stock_min must be numeric"
                            )
                            continue
                        if cantidad < 0 or precio < 0 or stock_min < 0:
                            errors.append(f"Row {row_num}: numeric values must be non-negative")
                            continue

                        crear_producto_fn(
                            codigo=codigo,
                            nombre=nombre,
                            cantidad=cantidad,
                            precio=precio,
                            categoria=row.get("categoria", ""),
                            descripcion=row.get("descripcion", ""),
                            stock_min=stock_min,
                        )
                        success += 1
                    except Exception as row_err:
                        errors.append(f"Row {row_num}: {row_err}")

            logger.info(f"CSV import finished: {success} success, {len(errors)} errors")
            return success, errors
        except InventarioException:
            raise
        except Exception as e:
            logger.error(f"Error importing CSV: {e}")
            raise InventarioException(f"CSV import failed: {e}")

    @staticmethod
    def import_from_xlsx(
        filepath: str,
        crear_producto_fn: Callable[..., dict],
    ) -> tuple[int, list[str]]:
        """Import products from an XLSX file (openpyxl).

        Expected header (case-insensitive, row 1): codigo, nombre, cantidad,
        precio, categoria (optional), descripcion (optional),
        stock_min (optional). Same shape as the CSV importer.

        Args:
            filepath: Path to the .xlsx file.
            crear_producto_fn: Callable that receives kwargs and creates the
                product in the DB. Errors per row are caught and surfaced.

        Returns:
            Tuple of (success_count, list_of_error_messages_per_row).
        """
        path = Path(filepath)
        if not path.exists():
            raise InventarioException(f"XLSX file not found: {filepath}")

        success = 0
        errors: list[str] = []
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise InventarioException("XLSX appears to be empty")

            # Read header row and build a normalized {lowercase: column_index}
            # map so the user can use any capitalization in their file.
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                raise InventarioException("XLSX has no rows")

            if not header_row:
                raise InventarioException("XLSX has no header row")

            field_map: dict[str, int] = {}
            for idx, cell in enumerate(header_row):
                if cell is None:
                    continue
                key = str(cell).lower().strip()
                if key:
                    field_map[key] = idx

            required = ("codigo", "nombre")
            for r in required:
                if r not in field_map:
                    raise InventarioException(f"Missing required column: {r}")

            def _get(row, key):
                idx = field_map.get(key)
                if idx is None or idx >= len(row):
                    return ""
                v = row[idx]
                return "" if v is None else str(v).strip()

            for row_num, row in enumerate(rows, start=2):
                if row is None:
                    continue
                # Skip fully-empty rows
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                try:
                    codigo = _get(row, "codigo")
                    nombre = _get(row, "nombre")
                    if not codigo or not nombre:
                        errors.append(f"Row {row_num}: codigo and nombre are required")
                        continue
                    try:
                        cantidad = int(_get(row, "cantidad") or "0")
                        precio = float(_get(row, "precio") or "0")
                        stock_min = int(_get(row, "stock_min") or "0")
                    except ValueError:
                        errors.append(f"Row {row_num}: cantidad/precio/stock_min must be numeric")
                        continue
                    if cantidad < 0 or precio < 0 or stock_min < 0:
                        errors.append(f"Row {row_num}: numeric values must be non-negative")
                        continue

                    crear_producto_fn(
                        codigo=codigo,
                        nombre=nombre,
                        cantidad=cantidad,
                        precio=precio,
                        categoria=_get(row, "categoria"),
                        descripcion=_get(row, "descripcion"),
                        stock_min=stock_min,
                    )
                    success += 1
                except Exception as row_err:
                    errors.append(f"Row {row_num}: {row_err}")

            wb.close()
            logger.info(f"XLSX import finished: {success} success, {len(errors)} errors")
            return success, errors
        except InventarioException:
            raise
        except Exception as e:
            logger.error(f"Error importing XLSX: {e}")
            raise InventarioException(f"XLSX import failed: {e}")
